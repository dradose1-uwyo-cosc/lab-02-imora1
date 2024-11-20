#Isabell Mora
#UW COSC 1010
#11-19-24
#HW05
#sources: , people worked with:
import openpyxl
from openpyxl.styles import PatternFill
import string

wb = openpyxl.Workbook()
sheet = wb.active
colors = {'B':'0000FF', 'R':'FF0000', 'Y':'FFFFFF','G':'008000','P':'FF69B4'}
cells = {
    'A1': 'R', 'B1': 'R', 'C1': 'R', 'D1': 'R', 'E1': 'R',
    'A2': 'R', 'B2': 'P', 'C2': 'Y', 'D2': 'P', 'E2': 'R',
    'A3': 'R', 'B3': 'Y', 'C3': 'Y', 'D3': 'Y', 'E3': 'R',
    'A4': 'R', 'B4': 'Y', 'C4': 'Y', 'D4': 'Y', 'E4': 'R',
    'A5': 'R', 'B5': 'Y', 'C5': 'Y', 'D5': 'Y', 'E5': 'R',
    'A6': 'R', 'B6': 'P', 'C6': 'Y', 'D6': 'P', 'E6': 'R',
    'A7': 'R', 'B7': 'R', 'C7': 'R', 'D7': 'R', 'E7': 'R',
    'B0': 'P', 'C0': 'P', 'D0': 'P', 'C8': 'P', 'D8': 'P', 
    'C9': 'P','A8': 'G', 'A9': 'G', 'A10': 'G','B8': 'G', 
    'B9': 'G','C8': 'G', 'C9': 'G','D9': 'G', 'D10': 'G',
    'F8': 'G', 'F9': 'G', 'F10': 'G','G8': 'G', 'G9': 'G',
    'F7': 'G',
}
c_sky = PatternFill(fill_type='solid', start_color='0000FF', end_color='0000FF')
fill = {}
for coord, color_key in cells.items():
    color = colors[color_key]
    fill[coord] = PatternFill(start_color=color, end_color=color, fill_type='solid')

for chr in string.ascii_uppercase[:12]:
    sheet.column_dimensions[chr].width = 5

for i in range(1,11):
    sheet.row_dimensions[i].height = 20

for chr in string.ascii_uppercase[:12]:
    for i in range(1,11):
        coord = chr +str(i)
        if coord in cells:
            sheet[coord].fill = fill[coord]
        else:
            sheet[coord].fill = c_sky
wb.save("flower_design.xlsx")









        
